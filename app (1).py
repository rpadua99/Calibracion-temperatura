"""
App de captura de calibración — Reporte de Temperatura
=======================================================

Flujo único: formulario guiado -> llena el Excel real (respetando todas
sus fórmulas) -> recalcula -> genera PDF listo para firmar.

Cómo correrla:
    pip install streamlit openpyxl
    streamlit run app.py

Requiere LibreOffice instalado en el sistema (`soffice`) para exportar el PDF.
Coloca el archivo de plantilla real junto a este script con el nombre
'template.xlsx' (o cambia TEMPLATE_PATH abajo).
"""

import io
import json
import subprocess
import tempfile
import datetime
from pathlib import Path

import streamlit as st
import openpyxl
from openpyxl.worksheet.header_footer import _HeaderFooterPart

# ---------------------------------------------------------------------------
TEMPLATE_PATH = Path(__file__).parent / "template.xlsx"
LOCALE_ES_MX = "[$-080A]"
# Fila de la tabla de resultados en "Reporte (2)" para cada punto (1 a 5)
POINT_ROWS = {1: 34, 2: 35, 3: 36, 4: 37, 5: 38}

PATRONES = [
    {"no": 1, "codigo": "02-0G-00627", "descripcion": "Sensor de temperatura RTD Patrón"},
    {"no": 2, "codigo": "02-02-03948", "descripcion": "Termómetro digital"},
    {"no": 3, "codigo": "02-0G-01389", "descripcion": "Sensor de temperatura RTD Patrón"},
    {"no": 4, "codigo": "02-0G-00955", "descripcion": "Sensor de temperatura"},
    {"no": 5, "codigo": "02-0G-00625", "descripcion": "Sensor de temperatura RTD Patrón"},
    {"no": 6, "codigo": "02-0G-00626 REF 1", "descripcion": "Referencia baja temperatura"},
    {"no": 7, "codigo": "02-0G-00626 REF 2", "descripcion": "Referencia baja temperatura"},
    {"no": 8, "codigo": "02-0G-00626 (2)", "descripcion": "Referencia baja temperatura"},
    {"no": 9, "codigo": "02-0G-00626 (4)", "descripcion": "Referencia baja temperatura"},
    {"no": 10, "codigo": "02-02-05267", "descripcion": "Sensor de temperatura"},
    {"no": 11, "codigo": "02-0G-01781", "descripcion": "Sensor de temperatura"},
    {"no": 12, "codigo": "02-02-05858", "descripcion": "Calibrador de temperatura"},
    {"no": 13, "codigo": "02-02-03967 (1)", "descripcion": "Calibrador de procesos tipo K (1)"},
    {"no": 14, "codigo": "02-02-03967 (2)", "descripcion": "Calibrador de procesos tipo K (2)"},
    {"no": 15, "codigo": "02-02-04961", "descripcion": "Sensor de temperatura y humedad relativa"},
    {"no": 16, "codigo": "02-02-04962", "descripcion": "Sensor de temperatura y humedad relativa"},
]

INGENIEROS = [
    "Daniel Guadarrama Mejía",
    "Brenda Torres Salazar",
    "Marco Antonio Esquivel Balderas",
    "Andrés López Godínez",
    "Ricardo Israel Padua Téllez",
]


@st.cache_data
def load_catalog():
    """Catálogo de instrumentos de temperatura, extraído una vez del template."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)
    ws = wb["datos crudos"]
    items = []
    kws = ["temperatura", "termómetro", "termohigrómetro", "termopar", "rtd", "termocupla"]
    for r in range(3, ws.max_row + 1):
        code = ws.cell(r, 36).value  # AJ
        desc = ws.cell(r, 37).value  # AK
        if code and desc and any(k in desc.lower() for k in kws):
            items.append((code, desc))
    return items


def fill_workbook(data: dict) -> bytes:
    """Llena SOLO las celdas de captura del template real y devuelve los bytes."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)
    ws = wb["datos crudos"]

    ws["N1"] = data["fecha"]
    ws["N3"] = data["anio"]
    ws["O3"] = data["consecutivo"]
    ws["N4"] = data["realizo"]
    ws["N5"] = data["codigo_instrumento"]
    ws["S1"] = data["patron_no"]

    ws["N11"] = data["l_inferior"]
    ws["O11"] = data["unidad"]
    ws["N12"] = data["l_superior"]
    ws["N13"] = data["division_minima"]
    ws["N14"] = data["factor_resolucion"]
    ws["N15"] = data["puntos_a_calibrar"]
    ws["N16"] = data["repeticiones"]
    ws["N17"] = data["tolerancia"]
    ws["O24"] = data["num_decimales"]

    ws["R23"] = data["temp_inicial"]
    ws["S23"] = data["temp_final"]
    ws["R24"] = data["hr_inicial"]
    ws["S24"] = data["hr_final"]

    for col, val in zip(["N27", "O27", "P27", "R27", "S27"], data["puntos_nominales"]):
        if val is not None:
            ws[col] = val

    bloques = [(4, "V", "W"), (4, "AA", "AB"), (4, "AF", "AG"), (12, "V", "W"), (12, "AA", "AB")]
    for (fila0, col_ibc, col_pat), punto in zip(bloques, data["lecturas"]):
        for i in range(5):
            fila = fila0 + i
            if i < len(punto["ibc"]) and punto["ibc"][i] is not None:
                ws[f"{col_ibc}{fila}"] = punto["ibc"][i]
            if i < len(punto["patron"]) and punto["patron"][i] is not None:
                ws[f"{col_pat}{fila}"] = punto["patron"][i]

    # --- Ocultar filas de puntos no usados en "Reporte (2)" (evita #DIV/0! visibles) ---
    rep = wb["Reporte (2)"]
    puntos = data["puntos_a_calibrar"]
    for punto_no, fila in POINT_ROWS.items():
        rep.row_dimensions[fila].hidden = punto_no > puntos

    # --- Fechas en español (independiente del locale de quien recalcule/imprima) ---
    for coord in ["L5", "L22", "K23", "K42", "K43"]:
        cell = rep[coord]
        if cell.number_format and "mmm" in cell.number_format.lower() and not cell.number_format.startswith("[$-"):
            cell.number_format = LOCALE_ES_MX + cell.number_format

    apply_clean_footer(rep)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def apply_clean_footer(ws):
    """Pie de página limpio (igual al original, sin la imagen &G redundante
    ni los saltos de línea rotos que introduce openpyxl al reescribir)."""
    ws.oddFooter.left = _HeaderFooterPart(text="FORM-000039461, Ed. 1", font="Optima,Normal", size=8)
    ws.oddFooter.center = _HeaderFooterPart(text="Código anterior: NA", font="Optima,Normal", size=8)
    ws.oddFooter.right = _HeaderFooterPart(text="Página &P de &N", font="Optima,Normal", size=8)


RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def recalculate_inplace(path: Path, timeout: int = 90):
    """Recalcula TODAS las fórmulas del libro con LibreOffice y guarda los
    valores calculados en el mismo archivo. Usa una macro (calculateAll +
    store) porque una simple conversión de formato no fuerza el recálculo.
    Autosuficiente: no depende de ningún archivo fuera de este script."""
    with tempfile.TemporaryDirectory() as profile_dir:
        profile_dir = Path(profile_dir)
        profile_url = profile_dir.as_uri()

        subprocess.run(
            ["soffice", "--headless", "--terminate_after_init",
             f"-env:UserInstallation={profile_url}"],
            capture_output=True, timeout=timeout,
        )
        macro_dir = profile_dir / "user" / "basic" / "Standard"
        macro_dir.mkdir(parents=True, exist_ok=True)
        (macro_dir / "Module1.xba").write_text(RECALCULATE_MACRO)

        subprocess.run(
            ["soffice", "--headless", "--norestore",
             f"-env:UserInstallation={profile_url}",
             "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
             str(path)],
            capture_output=True, timeout=timeout, check=True, text=True,
        )


def recalculate_and_pdf(xlsx_bytes: bytes):
    """Recalcula con LibreOffice (forzando calculateAll) y devuelve (xlsx_recalculado, pdf) como bytes."""
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx_path = tmp / "reporte.xlsx"
        xlsx_path.write_bytes(xlsx_bytes)

        # Recalcular todas las fórmulas del libro (macro calculateAll + store,
        # que es la única forma confiable de que LibreOffice cachee los valores)
        recalculate_inplace(xlsx_path, timeout=90)
        # Nota: el pie de página ya se dejó "limpio" en fill_workbook() ANTES de
        # llegar aquí. LibreOffice preserva un pie de página simple correctamente
        # al recalcular; el problema solo aparece si se le pasa el pie de página
        # original (con imagen incrustada y saltos de línea), así que el orden
        # de los pasos importa.
        recalculated = xlsx_path.read_bytes()

        # Ocultar hojas auxiliares y exportar solo "Reporte (2)" a PDF
        wb = openpyxl.load_workbook(xlsx_path)
        wb["datos crudos"].sheet_state = "hidden"
        wb["Patrones"].sheet_state = "hidden"
        wb.active = wb.sheetnames.index("Reporte (2)")
        apply_clean_footer(wb["Reporte (2)"])  # una vez más, por si acaso
        pdf_source = tmp / "reporte_pdf.xlsx"
        wb.save(pdf_source)

        subprocess.run(
            ["soffice", "--headless", "--convert-to", "pdf", "--outdir", str(tmp), str(pdf_source)],
            check=True, capture_output=True, timeout=90,
        )
        pdf_bytes = (tmp / "reporte_pdf.pdf").read_bytes()

    return recalculated, pdf_bytes


# ---------------------------------------------------------------------------
st.set_page_config(page_title="Captura de calibración · Temperatura", layout="centered")
st.title("🌡️ Captura de calibración — Temperatura")
st.caption("Formulario guiado → Excel real recalculado → PDF listo para firmar")

if not TEMPLATE_PATH.exists():
    st.error(f"No encuentro la plantilla en {TEMPLATE_PATH}. Colócala junto a este script como 'template.xlsx'.")
    st.stop()

catalog = load_catalog()

with st.form("captura"):
    st.subheader("1. Datos generales")
    c1, c2 = st.columns(2)
    fecha = c1.date_input("Fecha de calibración", datetime.date.today())
    consecutivo = c2.number_input("Consecutivo", min_value=1, step=1)
    realizo = st.selectbox("Realizó", INGENIEROS)

    st.subheader("2. Instrumento a calibrar")
    codigo_instrumento = st.selectbox(
        "Código", options=[c for c, _ in catalog],
        format_func=lambda c: f"{c} — {dict(catalog)[c]}",
    )
    c3, c4, c5 = st.columns(3)
    l_inferior = c3.number_input("Límite inferior", value=0.0)
    l_superior = c4.number_input("Límite superior", value=0.0)
    unidad = c5.selectbox("Unidad", ["°C", "°F", "°R"])
    c6, c7, c8 = st.columns(3)
    division_minima = c6.number_input("División mínima", value=0.1)
    tolerancia = c7.number_input("Tolerancia (±)", value=0.5)
    num_decimales = c8.number_input("Decimales", value=2, min_value=0, max_value=4, step=1)
    factor_resolucion = st.number_input("Factor de resolución", value=1)

    st.subheader("3. Patrón de referencia")
    patron_no = st.selectbox(
        "Patrón", options=[p["no"] for p in PATRONES],
        format_func=lambda no: f"{no}. {next(p['codigo'] for p in PATRONES if p['no']==no)} — "
                                f"{next(p['descripcion'] for p in PATRONES if p['no']==no)}",
    )

    st.subheader("4. Puntos de calibración")
    c9, c10 = st.columns(2)
    puntos_a_calibrar = c9.selectbox("Puntos a calibrar", [1, 2, 3, 4, 5], index=2)
    repeticiones = c10.selectbox("Repeticiones por punto", [1, 2, 3, 4, 5], index=4)

    puntos_nominales = []
    cols_pn = st.columns(5)
    for i in range(5):
        disabled = i >= puntos_a_calibrar
        val = cols_pn[i].number_input(f"Nominal Pto {i+1}", value=0.0, disabled=disabled, key=f"nom{i}")
        puntos_nominales.append(val if not disabled else None)

    st.subheader("5. Condiciones ambientales")
    c11, c12, c13, c14 = st.columns(4)
    temp_inicial = c11.number_input("Temp. inicial", value=20.0)
    temp_final = c12.number_input("Temp. final", value=20.0)
    hr_inicial = c13.number_input("%HR inicial", value=45.0)
    hr_final = c14.number_input("%HR final", value=45.0)

    st.subheader("6. Lecturas")
    lecturas = []
    for p in range(puntos_a_calibrar):
        st.markdown(f"**Punto {p + 1}**" + (f" · nominal {puntos_nominales[p]} {unidad}" if puntos_nominales[p] is not None else ""))
        ibc_vals, pat_vals = [], []
        cols = st.columns(repeticiones)
        for r in range(repeticiones):
            ibc_vals.append(cols[r].number_input(f"IBC #{r+1}", value=0.0, key=f"ibc_{p}_{r}"))
        cols2 = st.columns(repeticiones)
        for r in range(repeticiones):
            pat_vals.append(cols2[r].number_input(f"Patrón #{r+1}", value=0.0, key=f"pat_{p}_{r}"))
        lecturas.append({"ibc": ibc_vals, "patron": pat_vals})

    submitted = st.form_submit_button("Generar reporte (Excel + PDF)", use_container_width=True)

if submitted:
    data = {
        "fecha": fecha,
        "anio": int(str(fecha.year)[-2:]),
        "consecutivo": int(consecutivo),
        "realizo": realizo,
        "codigo_instrumento": codigo_instrumento,
        "patron_no": patron_no,
        "l_inferior": l_inferior,
        "l_superior": l_superior,
        "division_minima": division_minima,
        "factor_resolucion": factor_resolucion,
        "unidad": unidad,
        "puntos_a_calibrar": puntos_a_calibrar,
        "repeticiones": repeticiones,
        "tolerancia": tolerancia,
        "num_decimales": num_decimales,
        "temp_inicial": temp_inicial,
        "temp_final": temp_final,
        "hr_inicial": hr_inicial,
        "hr_final": hr_final,
        "puntos_nominales": puntos_nominales,
        "lecturas": lecturas,
    }

    with st.spinner("Llenando plantilla y recalculando fórmulas…"):
        xlsx_bytes = fill_workbook(data)
        try:
            xlsx_final, pdf_final = recalculate_and_pdf(xlsx_bytes)
        except FileNotFoundError:
            st.error("No se encontró LibreOffice ('soffice') en este sistema. "
                      "Instálalo para habilitar el recálculo y la exportación a PDF.")
            st.stop()
        except subprocess.CalledProcessError as e:
            st.error(f"Error al recalcular/exportar: {e}")
            st.stop()

    st.success("Reporte generado ✅")
    c15, c16 = st.columns(2)
    c15.download_button(
        "⬇️ Descargar Excel", data=xlsx_final,
        file_name=f"reporte_{codigo_instrumento}_{consecutivo}.xlsx",
        use_container_width=True,
    )
    c16.download_button(
        "⬇️ Descargar PDF", data=pdf_final,
        file_name=f"reporte_{codigo_instrumento}_{consecutivo}.pdf",
        use_container_width=True,
    )
